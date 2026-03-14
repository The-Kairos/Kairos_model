"""Compare LLM scene descriptions across log files and export to Excel."""

from __future__ import annotations

import glob
import json
import os
from collections import defaultdict

import pandas as pd
from openpyxl.styles import Alignment

DEFAULT_LOG_DIR = "./logs/runs"
DEFAULT_OUTPUT_EXCEL = "./logs/reports/llm_descriptions_comparisons.xlsx"

SHARED_COLS = [
    "frame_captions",
    "yolo_detections",
    "audio_natural",
    "audio_speech",
]

COLUMN_WIDTH = 60
CHARS_PER_LINE = 60
LINE_HEIGHT = 15


def format_yolo(yolo_dict: object) -> str:
    """Convert YOLO detection dict into human-readable text."""
    if not isinstance(yolo_dict, dict):
        return ""

    lines: list[str] = []
    for frame_idx, detections in yolo_dict.items():
        if not detections:
            continue
        lines.append(f"Frame {frame_idx}:")
        for det in detections:
            label = det.get("label", "unknown")
            conf = det.get("confidence")
            if conf is not None:
                conf = round(conf, 2)
                lines.append(f"- {label} ({conf})")
            else:
                lines.append(f"- {label}")
        lines.append("")

    return "\n".join(lines).strip()


def main() -> None:
    """Entry point for the description-comparison CLI."""
    log_dir = DEFAULT_LOG_DIR
    output_excel = DEFAULT_OUTPUT_EXCEL

    os.makedirs(os.path.dirname(output_excel), exist_ok=True)

    files = glob.glob(os.path.join(log_dir, "*.json"))

    groups: dict[str, list[str]] = defaultdict(list)
    for filepath in files:
        prefix = os.path.basename(filepath).split("_")[0]
        groups[prefix].append(filepath)

    wrap_top_left = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        for sheet_name, group_files in groups.items():
            data: dict[str, list[str]] = {}
            max_scenes = 0

            for filepath in group_files:
                with open(filepath, "r", encoding="utf-8") as f:
                    log = json.load(f)

                scenes = log.get("scenes", [])
                descriptions = [
                    scene.get("llm_scene_description", "") for scene in scenes
                ]

                col_name = os.path.basename(filepath)
                data[col_name] = descriptions
                max_scenes = max(max_scenes, len(descriptions))

            for col in data:
                data[col] += [""] * (max_scenes - len(data[col]))

            with open(group_files[-1], "r", encoding="utf-8") as f:
                shared_log = json.load(f)

            shared_scenes = shared_log.get("scenes", [])

            for col in SHARED_COLS:
                values: list[str] = []
                for i in range(max_scenes):
                    if i < len(shared_scenes):
                        val = shared_scenes[i].get(col, "")
                        if col == "yolo_detections":
                            val = format_yolo(val)
                        elif isinstance(val, (list, dict)):
                            val = json.dumps(val, ensure_ascii=False, indent=2)
                    else:
                        val = ""
                    values.append(val)
                data[col] = values

            df = pd.DataFrame(data)
            sheet = sheet_name[:31]
            df.to_excel(writer, sheet_name=sheet, index=False)

            ws = writer.sheets[sheet]

            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                ws.column_dimensions[col_letter].width = COLUMN_WIDTH

            for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=1,
                max_col=ws.max_column,
            ):
                for cell in row:
                    cell.alignment = wrap_top_left

            for row in ws.iter_rows(
                min_row=2,
                max_row=ws.max_row,
                min_col=1,
                max_col=3,
            ):
                max_lines = 1
                for cell in row:
                    if cell.value:
                        lines = max(
                            1,
                            len(str(cell.value)) // CHARS_PER_LINE,
                        )
                        max_lines = max(max_lines, lines)

                ws.row_dimensions[row[0].row].height = max(
                    LINE_HEIGHT, max_lines * LINE_HEIGHT
                )

    print(f"Excel written to: {output_excel}")


if __name__ == "__main__":
    main()
