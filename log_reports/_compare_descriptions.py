import json
import os
import glob
import pandas as pd
from collections import defaultdict
from openpyxl.styles import Alignment

# ----------------------------
# Config
# ----------------------------
LOG_DIR = "./logs"
OUTPUT_EXCEL = "./log_reports/llm_descriptions_comparisons.xlsx"

SHARED_COLS = [
    "frame_captions",
    "yolo_detections",
    "audio_natural",
    "audio_speech",
]

COLUMN_WIDTH = 60
CHARS_PER_LINE = 60
LINE_HEIGHT = 15

os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)

# ----------------------------
# Helpers
# ----------------------------
def format_yolo(yolo_dict):
    """
    Convert YOLO detections into human-readable text.
    """
    if not isinstance(yolo_dict, dict):
        return ""

    lines = []
    for frame_idx, detections in yolo_dict.items():
        if not detections:
            continue

        lines.append(f"Frame {frame_idx}:")
        for det in detections:
            label = det.get("label", "unknown")
            conf = det.get("confidence")
            if conf is not None:
                conf = round(conf, 2)
                lines.append(f"- {label} ({conf})")
            else:
                lines.append(f"- {label}")

        lines.append("")  # blank line between frames

    return "\n".join(lines).strip()

# ----------------------------
# Load JSON files
# ----------------------------
files = glob.glob(os.path.join(LOG_DIR, "*.json"))

groups = defaultdict(list)
for filepath in files:
    prefix = os.path.basename(filepath).split("_")[0]
    groups[prefix].append(filepath)

# ----------------------------
# Styles
# ----------------------------
wrap_top_left = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True,
)

# ----------------------------
# Write Excel
# ----------------------------
with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
    for sheet_name, group_files in groups.items():
        data = {}
        max_scenes = 0

        # ---- LLM description columns ----
        for filepath in group_files:
            with open(filepath, "r", encoding="utf-8") as f:
                log = json.load(f)

            scenes = log.get("scenes", [])
            descriptions = [
                scene.get("llm_scene_description", "")
                for scene in scenes
            ]

            col_name = os.path.basename(filepath)
            data[col_name] = descriptions
            max_scenes = max(max_scenes, len(descriptions))

        # Pad LLM columns
        for col in data:
            data[col] += [""] * (max_scenes - len(data[col]))

        # ---- Shared columns (from last file only) ----
        with open(group_files[-1], "r", encoding="utf-8") as f:
            shared_log = json.load(f)

        shared_scenes = shared_log.get("scenes", [])

        for col in SHARED_COLS:
            values = []
            for i in range(max_scenes):
                if i < len(shared_scenes):
                    val = shared_scenes[i].get(col, "")

                    if col == "yolo_detections":
                        val = format_yolo(val)

                    elif isinstance(val, (list, dict)):
                        val = json.dumps(val, ensure_ascii=False, indent=2)

                else:
                    val = ""

                values.append(val)

            data[col] = values

        # ---- Write sheet ----
        df = pd.DataFrame(data)
        sheet = sheet_name[:31]
        df.to_excel(writer, sheet_name=sheet, index=False)

        ws = writer.sheets[sheet]

        # ---- Column widths ----
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = COLUMN_WIDTH

        # ---- Wrap & align ALL cells ----
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                cell.alignment = wrap_top_left

        # ---- Adjust row height (based on first 3 columns) ----
        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=1,
            max_col=3,
        ):
            max_lines = 1
            for cell in row:
                if cell.value:
                    lines = max(1, len(str(cell.value)) // CHARS_PER_LINE)
                    max_lines = max(max_lines, lines)

            ws.row_dimensions[row[0].row].height = max(
                LINE_HEIGHT, max_lines * LINE_HEIGHT
            )

print(f"✅ Excel written to: {OUTPUT_EXCEL}")
